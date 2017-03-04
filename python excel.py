from openpyxl import Workbook
wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
import datetime
ws['A3'] = datetime.datetime.now()

# another method to assign value to cell by row & column address
d = ws.cell(row=10, column=2, value=100)

# Rename the sheet title
ws.title = "Data"

# Save the file
wb.save("sample.xlsx")

##########################################################################

#You can also create new worksheets by using 
ws = wb.create_sheet("first sheet", 0) # insert at first sheet tab
ws = wb.create_sheet("last sheet") # insert at last sheet tab (default)
# Save the file
wb.save("sample.xlsx")

##########################################################################
#Review the names of all worksheets of the workbook
print(wb.sheetnames)

##########################################################################

# create new sheet and access column c
ws = wb.create_sheet("accessing cell")
colC = ws['C']
print (colC)

# access column range C:D
col_range = ws['C:D']
print (col_range)

# access row 10
row10 = ws[10]
print(row10)

# access row range 5:10
row_range = ws[5:10]
print (row_range)

#obtain range of cells (rows & columns)
cell_range = ws['A1':'C2']
print (cell_range)

wb.save("sample.xlsx")



